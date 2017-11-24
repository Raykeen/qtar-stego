import openpyxl
from openpyxl.utils import (get_column_letter)

first_col = 1
first_row = 1


def save_de_results(file, def_params, new_params, def_metrics, new_metrics):
    try:
        workbook = openpyxl.load_workbook(file, guess_types=True)
    except:
        workbook = openpyxl.Workbook()
        workbook.guess_types = True

    issue = def_params["issue"]

    sheet_name = "issue " + str(issue) + " " +\
                 ('pm' if def_params['pm_mode'] else '') +\
                 ('cf' if def_params['cf_mode'] else '') +\
                 ('wmdct' if def_params['wmdct_mode'] else '')

    def_params = prepare_params(def_params)
    new_params = prepare_params(new_params)
    def_metrics = prepare_params(def_metrics)
    new_metrics = prepare_params(new_metrics)

    try:
        sheet = workbook.get_sheet_by_name(sheet_name)
    except:
        sheet = workbook.create_sheet(sheet_name)
        headers = (*def_params["headers"], *def_metrics["headers"], *new_params["headers"], *new_metrics["headers"])
        past_list_in_row(sheet, first_col, first_row, headers)

    row = first_row + 1
    container = def_params["values"][0]
    watermark = def_params["values"][1]
    while True:
        container_cell_val = sheet[cell(first_col, row)].value
        watermark_cell_val = sheet[cell(first_col + 1, row)].value
        if container_cell_val is None:
            break
        if container_cell_val == container and watermark_cell_val == watermark:
            break
        row += 1

    values = (*def_params["values"], *def_metrics["values"], *new_params["values"], *new_metrics["values"])
    past_list_in_row(sheet, first_col, row, values)
    workbook.save(file)


def cell(col, row):
    return get_column_letter(col) + str(row)


def cells_range(col1, row1, col2, row2):
    return cell(col1, row1) + ":" + cell(col2, row2)


def prepare_params(params):
    headers = []
    values = []
    if "container" in params:
        headers.append("Контейнер")
        values.append(params['container'].split("\\")[-1].split(".")[0])
    if "watermark" in params:
        headers.append("Секретное изображение")
        values.append(params['watermark'].split("\\")[-1].split(".")[0])
    if "homogeneity_threshold" in params:
        th = params["homogeneity_threshold"]
        if isinstance(th, (tuple, list)):
            headers.extend("th" + str(i) for i in range(1, len(th) + 1))
            values.extend(th)
        else:
            headers.append("th")
            values.append(th)
    if "min_block_size" in params:
        headers.append("min_b, px")
        values.append(params["min_block_size"])
    if "max_block_size" in params:
        headers.append("max_b, px")
        values.append(params["max_block_size"])
    if "quant_power" in params:
        headers.append("q")
        values.append(params["quant_power"])
    if "ch_scale" in params:
        headers.append("k")
        values.append(params["ch_scale"])
    if "offset" in params:
        headers.extend(("x, px", "y, px"))
        values.extend((params["offset"][0], params["offset"][1]))
    if not ('cf_mode' in params and params['cf_mode'] is False) and 'cf_grid_size' in params:
        headers.append('cf')
        values.append(params['cf_grid_size'])
    if not ('wmdct_mode' in params and params['wmdct_mode'] is False):
        if 'wmdct_block_size' in params:
            headers.append('v, px')
            values.append(params['wmdct_block_size'])
        if 'wmdct_scale' in params:
            headers.append('sk')
            values.append(params['wmdct_scale'])
    if "container psnr" in params:
        headers.append("PSNR_C, db")
        values.append(params["container psnr"])
    if "container ssim" in params:
        headers.append("SSIM_C")
        values.append(params["container ssim"])
    if "watermark psnr" in params:
        headers.append("PSNR_W, db")
        values.append(params["watermark psnr"])
    if "watermark ssim" in params:
        headers.append("SSIM_W")
        values.append(params["watermark ssim"])
    if "watermark bcr" in params:
        headers.append("BCR")
        values.append(params["watermark bcr"])
    if "container bpp" in params:
        headers.append("BPP")
        values.append(params["container bpp"])
    if "key size" in params:
        headers.append("Размер ключа, байт")
        values.append(params["key size"])
    return {
        "headers": headers,
        "values": values
    }


def past_list_in_row(sheet, col, row, list):
    for item in list:
        sheet[cell(col, row)] = item
        col += 1
    return col


def write_table(sheet, table):
    for row in table:
        sheet.append(row)

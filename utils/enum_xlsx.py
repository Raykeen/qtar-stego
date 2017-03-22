import openpyxl
import string
from os import listdir
from re import match


try:
    workbook = openpyxl.load_workbook('xls\\enum.xlsx', guess_types=True)
except:
    workbook = openpyxl.Workbook()
    workbook.guess_types = True

logs = listdir('logs')
enum_logs = [log for log in logs if match("enum_[_.\w]*.log", log)]

for log in enum_logs:
    log_file = open('logs\\'+log)
    name = log_file.readline().replace('\n', '')
    try:
        sheet = workbook.get_sheet_by_name(name)
    except:
        sheet = workbook.create_sheet(name)
        headers = ["x", "y", "psnr", "bcr", "bpp"]
        for col, val in zip(string.ascii_uppercase[:5], headers):
            sheet[col + '1'] = val

    row = 1
    for line in log_file.readlines():
        line = line.replace('\n', '')
        values = line.split(' ')
        row += 1
        for col, val in zip(string.ascii_uppercase[:5], values):
            sheet[col + str(row)] = val


workbook.save('xls\\enum.xlsx')

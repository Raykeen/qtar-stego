import openpyxl
import string
from os import listdir
from re import match
from openpyxl.utils import (get_column_letter)


try:
    workbook = openpyxl.load_workbook('xls\\robustness.xlsx', guess_types=True)
except:
    workbook = openpyxl.Workbook()
    workbook.guess_types = True

logs = listdir('logs')
enum_logs = [log for log in logs if match("robustness_[_.\w]*.log", log)]

try:
    sheet = workbook.get_sheet_by_name("robustness")
except:
    sheet = workbook.create_sheet("robustness")
    headers = ["Контейнер", "Секретное изобр."]
    for col, val in zip(range(1, 3), headers):
        sheet[get_column_letter(col) + '1'] = val


def set_to(col, row, value):
    sheet[get_column_letter(col) + str(row)] = value

row = 3
for log in enum_logs:
    log_file = open('logs\\'+log)
    for line in log_file:
        line = line.replace('\n', '')
        if match("[a-zA-Z\\\\_.]* in [a-zA-Z\\\\_.]*", line):
            wm, img = line.split(" in ")
            wm = wm.split('\\')[-1].split('.')[0]
            img = img.split('\\')[-1].split('.')[0]

            for col, val in zip(range(1, 3), [img, wm]):
                set_to(col, row, val)

            col = 3
            while True:
                line = log_file.readline().replace('\n', '')
                if match("_"*40, line):
                    break
                if match("FILTER: ", line):
                    filter_name = line.replace("FILTER: ", "")

                    while True:
                        line = log_file.readline().replace('\n', '')
                        if match("## ", line):
                            break

                    psnr, bcr = line.split(' ')[1:]
                    if row == 3:
                        set_to(col, 1, filter_name)
                        sheet.merge_cells(get_column_letter(col) + str(1) + ":" +
                                           get_column_letter(col + 1) + str(1))
                        set_to(col, 2, "psnr")
                        set_to(col + 1, 2, "bcr")

                    set_to(col, row, psnr)
                    set_to(col + 1, row, bcr)
                    col += 2

            row += 1


workbook.save('xls\\robustness.xlsx')

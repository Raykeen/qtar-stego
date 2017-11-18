from copy import copy
from itertools import product
import math

import openpyxl
from numpy import arange

from PIL import Image

from qtar.cli import embed
from qtar.core.argparser import create_argpaser
from qtar.utils.xlsx import past_list_in_row
from qtar.core.qtar import DEFAULT_PARAMS
from qtar.utils import extract_filename
from qtar.experiments import PARAMS_NAMES, METRICS_NAMES


def main():
    argparser = create_argpaser()
    argparser.add_argument('-rc', '--container_size', metavar='container_size',
                           type=int, nargs=2, default=None,
                           help='resize container image')
    argparser.add_argument('-rw', '--watermark_size', metavar='watermark_size',
                           type=int, nargs=2, default=None,
                           help='resize watermark')
    argparser.add_argument('-xls', '--xls_path', metavar='path',
                           type=str, default='xls/enumerations.xlsx',
                           help='save results to xls file')
    argparser.add_argument('-p', '--param', metavar='name',
                           type=str, default='quant_power',
                           help='one of: ' + ", ".join(DEFAULT_PARAMS.keys()))
    argparser.add_argument('-sp', '--step', metavar='value',
                           type=float, default=0.05,
                           help='change step')
    args = argparser.parse_args()
    params = vars(args)
    results = enumerate_quantization(params)
    to_xlsx(results, params)


def get_param_range(params):
    container_size = params['container_size'] if params['container_size'] is not None else Image.open(params['container']).size
    max_b = min(params['max_block_size'], min(container_size))
    min_b = params['min_block_size']
    step = params['step']
    int_step = int(math.ceil(step))
    param = params['param']

    return {
        'homogeneity_threshold': arange(0.01, 1.01, step),
        'min_block_size': map(lambda x: 2**x, range(3, int(math.log2(max_b)) + 1, int_step)),
        'max_block_size': map(lambda x: 2**x, range(int(math.log2(min_b)), int(math.log2(max_b)) + 1, int_step)),
        'quant_power': arange(0.1, 1, step),
        'cf_grid_size': map(lambda x: 2**x, range(0, int(math.log2(min_b)) + 1, int_step)),
        'ch_scale': arange(0.01, 256, step),
        'offset': product(range(0, container_size[0], int_step), range(0, container_size[1], int_step))
    }[param]


def enumerate_quantization(params):
    param = params["param"]
    results = []
    for p in get_param_range(params):
        testing_params = copy(params)
        testing_params[param] = p
        result = embed(testing_params)
        result = {METRICS_NAMES[key]: val for key, val in result.items()}
        result[PARAMS_NAMES[param]] = p
        results.append(result)

    return results


def to_xlsx(results, params):
    path = params["xls_path"]
    try:
        workbook = openpyxl.load_workbook(path, guess_types=True)
    except:
        workbook = openpyxl.Workbook()
        workbook.guess_types = True

    container_name = extract_filename(params["container"])

    sheet_name = "%s %s" % (PARAMS_NAMES[params['param']], container_name)
    if params['cf_mode']:
        sheet_name = 'cf ' + sheet_name
    if params['wmdct_mode']:
        sheet_name = 'wmdct ' + sheet_name
    if params['pm_mode']:
        sheet_name = 'pm ' + sheet_name

    try:
        sheet = workbook.get_sheet_by_name(sheet_name)
    except:
        sheet = workbook.create_sheet(sheet_name)

    past_list_in_row(sheet, 1, 1, results[0].keys())

    row = 2
    for result in results:
        past_list_in_row(sheet, 1, row, result.values())
        row += 1

    workbook.save(path)


if __name__ == "__main__":
    main()
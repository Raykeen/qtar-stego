import openpyxl
from os import listdir
from re import match
import string


try:
    workbook = openpyxl.load_workbook('xls\\de.xlsx', guess_types=True)
except:
    workbook = openpyxl.Workbook()
    workbook.guess_types = True

logs = listdir('logs')
enum_logs = [log for log in logs if match("de_[_.\w]*.log", log)]

rows = {}
for log in enum_logs:
    log_file = open('logs\\'+log)
    for line in log_file:
        line = line.replace('\n', '')
        if match("##\d##", line):
            name, th1, th2, th3, x, y = line.split(' ')
            name = name.replace('##', '')
            if not name in rows:
                rows[name] = 1
            while True:
                line = log_file.readline().replace('\n', '')
                if match("[a-zA-Z\\\\_.]* in [a-zA-Z\\\\_.]*", line):
                    break
            wm, img = line.split(" in ")
            wm = wm.split('\\')[-1].split('.')[0]
            img = img.split('\\')[-1].split('.')[0]
            while True:
                line = log_file.readline().replace('\n', '')
                if match("## ", line):
                    break
            psnr, bpp, bcr = line.split(' ')[1:]
            print("{0} {1} in {2} {3} {4} {5} {6} {7} {8} {9} {10}"
                  .format(name, wm, img, th1, th2, th3, x, y, psnr, bpp, bcr))
            try:
                sheet = workbook.get_sheet_by_name(name)
            except:
                sheet = workbook.create_sheet(name)
                headers = ["Контейнер", "Секретное изобр.", "th1", "th2", "th3", "x", "y", "psnr", "bcr", "bpp"]
                for col, val in zip(string.ascii_uppercase[:10], headers):
                    sheet[col+'1'] = val

            rows[name] += 1
            values = [img, wm, th1, th2, th3, x, y, psnr, bcr, bpp]
            for col, val in zip(string.ascii_uppercase[:10], values):
                sheet[col+str(rows[name])] = val

workbook.save('xls\\de.xlsx')
